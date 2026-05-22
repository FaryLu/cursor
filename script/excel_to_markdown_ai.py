from openpyxl import load_workbook

def excel_to_markdown_ai(excel_path: str, sheet_name: str = None):
    """
    读取 Excel 文件，转换成 Markdown 表格字符串，直接返回给AI（不生成文件）
    :param excel_path: Excel 文件路径（入参）
    :param sheet_name: 工作表名（默认读取第一个）
    :return: Markdown 表格字符串（AI直接用）
    """
    # 加载 Excel（只读模式，更快）
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    # 读取所有行数据
    rows = []
    for row in ws.iter_rows(values_only=True):
        # 过滤空行
        row_data = [str(cell).strip() if cell is not None else "" for cell in row]
        if any(cell.strip() for cell in row_data):
            rows.append(row_data)

    wb.close()

    # 转成 Markdown 表格
    md_lines = []
    for i, row in enumerate(rows):
        md_row = "| " + " | ".join(row) + " |"
        md_lines.append(md_row)
        # 表头下面加分隔线
        if i == 0:
            separator = "| " + " | ".join(["---"] * len(row)) + " |"
            md_lines.append(separator)

    return "\n".join(md_lines)